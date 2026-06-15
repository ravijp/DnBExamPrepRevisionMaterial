# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""Dump the DNB Radiodiagnosis tracker workbook to the terminal.

Reads DNB_Radiodiagnosis_Tracker.xlsx (binary, so it cannot be diffed in git)
and prints each sheet. The Tracker sheet additionally gets a per-category
roll-up: completeness, marks/session weight, and priority (frequency x gap).

Run from the repo root:
    uv run scripts/dump_tracker.py
    uv run scripts/dump_tracker.py --sheet Tracker      # one sheet only
    uv run scripts/dump_tracker.py --no-rollup          # raw rows only
"""
from __future__ import annotations

import argparse
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl

# Windows consoles default to cp1252 and choke on ★, →, etc. Force UTF-8.
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

WORKBOOK = Path(__file__).resolve().parent.parent / "DNB_Radiodiagnosis_Tracker.xlsx"


def _norm(s: str) -> str:
    return str(s or "").strip().lower().replace(" ", "").replace("_", "").replace("/", "").replace("%", "")


def find_col(headers: list[str], *candidates: str) -> int | None:
    """Return the index of the first header matching any candidate (fuzzy)."""
    nheaders = [_norm(h) for h in headers]
    for cand in candidates:
        nc = _norm(cand)
        for i, h in enumerate(nheaders):
            if nc and nc in h:
                return i
    return None


def dump_sheet(ws) -> list[list]:
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    print(f"\n{'=' * 70}\nSHEET: {ws.title}  ({ws.max_row} rows x {ws.max_column} cols)\n{'=' * 70}")
    for r in rows:
        cells = ["" if v is None else str(v) for v in r]
        if any(cells):
            print(" | ".join(cells))
    return rows


def rollup_tracker(rows: list[list]) -> None:
    if not rows:
        return
    # Locate the real header row (it contains both 'Paper' and 'Category');
    # earlier rows are the title/instruction banner.
    hdr_idx = 0
    for i, r in enumerate(rows):
        cells = [_norm(c) for c in r]
        if any("paper" == c for c in cells) and any("category" == c for c in cells):
            hdr_idx = i
            break
    headers = [str(h or "") for h in rows[hdr_idx]]
    data = rows[hdr_idx + 1:]

    i_cat = find_col(headers, "category")
    i_comp = find_col(headers, "completeness", "complete", "score")
    i_weight = find_col(headers, "catmarkssession", "marksession", "marks/session", "weight")
    i_topic = find_col(headers, "topic", "subcategory", "sub-category")

    if i_cat is None:
        print("\n[rollup] No 'Category' column found; skipping roll-up.")
        return

    cats: "OrderedDict[str, dict]" = OrderedDict()
    for r in data:
        cat = str(r[i_cat]).strip() if i_cat < len(r) and r[i_cat] is not None else ""
        if not cat:
            continue
        d = cats.setdefault(cat, {"n": 0, "comp": [], "weights": []})
        d["n"] += 1
        if i_comp is not None and i_comp < len(r):
            try:
                d["comp"].append(float(r[i_comp]))
            except (TypeError, ValueError):
                pass
        if i_weight is not None and i_weight < len(r):
            try:
                d["weights"].append(float(r[i_weight]))
            except (TypeError, ValueError):
                pass

    summary = []
    for cat, d in cats.items():
        avg_comp = sum(d["comp"]) / len(d["comp"]) if d["comp"] else None
        weight = max(d["weights"]) if d["weights"] else None  # weight is per-category, take representative
        gap = (10 - avg_comp) if avg_comp is not None else 10.0  # blank completeness => full gap
        priority = (weight * gap) if weight is not None else None
        summary.append((cat, d["n"], avg_comp, weight, gap, priority))

    summary.sort(key=lambda x: (x[5] is None, -(x[5] or 0)))

    print(f"\n{'=' * 70}\nPER-CATEGORY ROLL-UP  (priority = weight x gap; gap = 10 - completeness)\n{'=' * 70}")
    print(f"{'Category':<34}{'#topics':>8}{'compl':>8}{'weight':>9}{'gap':>7}{'priority':>10}")
    print("-" * 76)
    for cat, n, comp, weight, gap, prio in summary:
        c = f"{comp:.1f}" if comp is not None else "(blank)"
        w = f"{weight:.2f}" if weight is not None else "-"
        p = f"{prio:.2f}" if prio is not None else "-"
        print(f"{cat[:33]:<34}{n:>8}{c:>8}{w:>9}{gap:>7.1f}{p:>10}")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--sheet", help="Dump only this sheet")
    ap.add_argument("--no-rollup", action="store_true", help="Skip the Tracker roll-up")
    ap.add_argument("--workbook", default=str(WORKBOOK))
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    print(f"Workbook: {args.workbook}\nSheets: {wb.sheetnames}")

    for name in wb.sheetnames:
        if args.sheet and name != args.sheet:
            continue
        rows = dump_sheet(wb[name])
        if name.lower().startswith("tracker") and not args.no_rollup:
            rollup_tracker(rows)


if __name__ == "__main__":
    main()
